"""操作 chongli_ski_pass_orders_fy_2025-05-01_2026-04-30.xlsx 的「雪票列表」sheet：

1. 用「渠道订单号」（如 QJ_XP_260405_00001_ZF_02）匹配「年度雪票」的「订单号」（如
   QJ_XP_260405_00001）。匹配键 = 渠道订单号中 '_ZF_' 之前的部分。
2. 匹配不上的行（包括渠道订单号为空 / 不含 '_ZF_'）→ 整行字体设灰色（808080）。
3. 末尾追加一列「实际支付」：
   - 匹配上的 → 写入「年度雪票.订单结余」
   - 匹配不上的 → 空
4. 条件底色（仅匹配上的行参与）：
   - 订单状态='已完成' 且 实际支付 < LOW_THRESHOLD（默认 10）→ 红底 FFC7CE
   - 订单状态='已取消' 且 实际支付 > HIGH_THRESHOLD（默认 20）→ 黄底 FFEB9C

幂等：重跑会先重置数据区 Font/Fill 为默认、再按规则应用。「实际支付」列存在则覆盖。
其他 3 sheet（年度雪票 / 支付明细 / 支付流水）不动。
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

XLSX = r'D:\snowmeet\snowmeet_ai_doc\chongli_ski_pass_orders_fy_2025-05-01_2026-04-30.xlsx'
MAIN_SHEET = '年度雪票'
LIST_SHEET = '雪票列表'
NEW_COL = '实际支付'

LOW_THRESHOLD = 20    # 已完成 但实际支付 < 此值 → 红底
HIGH_THRESHOLD = 20   # 已取消 但实际支付 > 此值 → 黄底

GREY_FONT = Font(color='808080')
DEFAULT_FONT = Font()                          # 还原默认（含黑色字体）
DEFAULT_FILL = PatternFill(fill_type=None)     # 还原无填充
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')

HEADER_COLOR = '1F4E78'
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
HEADER_FILL = PatternFill('solid', fgColor=HEADER_COLOR)
CENTER = Alignment(horizontal='center', vertical='center')


def visual_width(s):
    return sum(2 if ord(ch) > 127 else 1 for ch in str(s))


def parse_order_code(channel_id):
    """渠道订单号 → snowmeet 订单号；无 _ZF_ / 空 → None"""
    if channel_id is None:
        return None
    s = str(channel_id).strip()
    if '_ZF_' not in s:
        return None
    return s.split('_ZF_')[0]


def main():
    if not os.path.exists(XLSX):
        raise SystemExit(f'xlsx 不存在: {XLSX}')

    print(f'打开 {XLSX}')
    wb = load_workbook(XLSX)
    for need in (MAIN_SHEET, LIST_SHEET):
        if need not in wb.sheetnames:
            raise SystemExit(f'缺 sheet「{need}」: {wb.sheetnames}')

    # ── 建主表 dict：订单号 → 订单结余 ──
    main_ws = wb[MAIN_SHEET]
    main_headers = [main_ws.cell(row=1, column=c).value for c in range(1, main_ws.max_column + 1)]
    if '订单号' not in main_headers or '订单结余' not in main_headers:
        raise SystemExit(f'年度雪票缺列: {main_headers}')
    # 「订单号」/「订单结余」可能在主表多列重复（段5 拷贝段1）；按用户语义优先用段5 的订单号列
    # 实际两列同值，取首个即可
    col_code = main_headers.index('订单号') + 1
    # 「订单结余」也可能出现两次（段1 中 14 + 段5 中 50），用「订单结余」+ 段5 取最右一个最稳妥
    # 实际两列同值取首
    code_to_balance = {}
    for r in range(2, main_ws.max_row + 1):
        code = main_ws.cell(row=r, column=col_code).value
        if not code:
            continue
        bal = main_ws.cell(row=r, column=main_headers.index('订单结余') + 1).value
        code_to_balance[str(code).strip()] = bal
    print(f'  年度雪票 dict 大小: {len(code_to_balance)}')

    # ── 操作雪票列表 ──
    ws = wb[LIST_SHEET]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if '渠道订单号' not in headers or '订单状态' not in headers:
        raise SystemExit(f'雪票列表缺列: {headers}')
    col_chan = headers.index('渠道订单号') + 1
    col_status = headers.index('订单状态') + 1
    print(f'  渠道订单号 列{col_chan}  订单状态 列{col_status}')

    # 定位/追加「实际支付」列
    if NEW_COL in headers:
        col_pay = headers.index(NEW_COL) + 1
        print(f'  「{NEW_COL}」已存在，覆盖列 {col_pay}')
    else:
        col_pay = ws.max_column + 1
        print(f'  追加「{NEW_COL}」到列 {col_pay}')
        cell = ws.cell(row=1, column=col_pay, value=NEW_COL)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # ── 数据区先全清样式（保表头）──
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).font = DEFAULT_FONT
            ws.cell(row=r, column=c).fill = DEFAULT_FILL

    stats = {'matched': 0, 'unmatched': 0, 'red': 0, 'yellow': 0,
             'finished_matched': 0, 'cancelled_matched': 0}

    for r in range(2, ws.max_row + 1):
        chan = ws.cell(row=r, column=col_chan).value
        status = ws.cell(row=r, column=col_status).value
        code = parse_order_code(chan)
        bal = code_to_balance.get(code) if code else None

        if bal is None:
            # 匹配不上：整行灰字 + 实际支付留空
            stats['unmatched'] += 1
            ws.cell(row=r, column=col_pay, value=None)
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).font = GREY_FONT
        else:
            stats['matched'] += 1
            pay_val = round(float(bal), 2) if isinstance(bal, (int, float)) else bal
            pay_cell = ws.cell(row=r, column=col_pay, value=pay_val)
            pay_cell.number_format = '0.00'

            # 条件底色（仅匹配上的）
            try:
                pay_num = float(pay_val) if pay_val is not None else None
            except (TypeError, ValueError):
                pay_num = None

            if status == '已完成':
                stats['finished_matched'] += 1
                if pay_num is not None and pay_num < LOW_THRESHOLD:
                    stats['red'] += 1
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = RED_FILL
            elif status == '已取消':
                stats['cancelled_matched'] += 1
                if pay_num is not None and pay_num > HIGH_THRESHOLD:
                    stats['yellow'] += 1
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = YELLOW_FILL

    # 设新列宽
    ws.column_dimensions[get_column_letter(col_pay)].width = max(
        visual_width(NEW_COL) + 2, 12)

    print()
    print(f'匹配: {stats["matched"]}  不匹配: {stats["unmatched"]}')
    print(f'匹配中 已完成: {stats["finished_matched"]}  已取消: {stats["cancelled_matched"]}')
    print(f'底色红（已完成+实付<{LOW_THRESHOLD}）: {stats["red"]}')
    print(f'底色黄（已取消+实付>{HIGH_THRESHOLD}）: {stats["yellow"]}')

    print(f'保存 {XLSX}')
    wb.save(XLSX)
    print('完成')


if __name__ == '__main__':
    main()
