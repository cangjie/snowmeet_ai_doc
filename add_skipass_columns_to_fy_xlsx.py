"""给 {prefix}_ski_pass_orders_fy_{start}_{end}.xlsx 的「年度雪票」sheet 末尾追加 4 列：
  雪票名称 (product_name) / 支付价格 (deal_price) / 结算价格 (ticket_price) / 取票时间 (card_member_pick_time)

数据源：ski_pass 表（valid=1，按 order_id 聚合到订单粒度）。
聚合口径（兜一单多票）：
  - product_name → 多张时分号 '; ' 连接，去重保序
  - deal_price / ticket_price → SUM
  - card_member_pick_time → MIN（最早一张取票时间）
  - 该订单无 ski_pass 行 → 4 列全空

用法：
  py add_skipass_columns_to_fy_xlsx.py --xlsx chongli_ski_pass_orders_fy_2025-05-01_2026-04-30.xlsx --shop 崇礼旗舰店
  py add_skipass_columns_to_fy_xlsx.py --xlsx nanshan_ski_pass_orders_fy_2025-05-01_2026-04-30.xlsx --shop 南山

幂等：若 4 列已存在（按列名匹配末尾若干列）则原地覆盖，不重复追加。
其他 sheet（支付明细 / 支付流水）不动。
"""
import argparse
import os
import sys

import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

SHEET = '年度雪票'
HEADER_COLOR = '1F4E78'

NEW_COLS = ['雪票名称', '支付价格', '结算价格', '取票时间']

CONN = ('DRIVER={ODBC Driver 18 for SQL Server};SERVER=tcp:100.28.143.19,1433;'
        'DATABASE=snowmeet_new;UID=claude;PWD=abcd123!@#;'
        'Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=30;')


def fetch_skipass_by_order(shop, start, end_excl):
    """返回 {order_code: (name_str, deal_sum, ticket_sum, min_pick_time)}，按订单聚合。"""
    cn = pyodbc.connect(CONN)
    cur = cn.cursor()
    # 拉行级数据 Python 端聚合（兼顾 product_name 字符串去重 + 时间 MIN）
    cur.execute("""SELECT o.code, sp.product_name, sp.deal_price, sp.ticket_price,
                          sp.card_member_pick_time
                   FROM ski_pass sp JOIN [order] o ON o.id = sp.order_id
                   WHERE o.shop = ? AND o.[type] = N'雪票'
                     AND o.biz_date >= ? AND o.biz_date < ?
                     AND o.valid = 1 AND o.code IS NOT NULL AND LTRIM(RTRIM(o.code)) <> N''
                     AND sp.valid = 1
                   ORDER BY o.id, sp.id""", shop, start, end_excl)
    by_code = {}
    for code, name, deal, ticket, pick in cur.fetchall():
        if code is None:
            continue
        key = code.strip()
        if key not in by_code:
            by_code[key] = {'names': [], 'deal': 0.0, 'ticket': 0.0, 'pick': None}
        rec = by_code[key]
        if name and name not in rec['names']:
            rec['names'].append(name)
        if deal is not None:
            rec['deal'] += float(deal)
        if ticket is not None:
            rec['ticket'] += float(ticket)
        if pick is not None and (rec['pick'] is None or pick < rec['pick']):
            rec['pick'] = pick
    cn.close()

    out = {}
    for code, rec in by_code.items():
        out[code] = (
            '; '.join(rec['names']) if rec['names'] else None,
            round(rec['deal'], 2) if rec['deal'] else None,
            round(rec['ticket'], 2) if rec['ticket'] else None,
            rec['pick'],
        )
    return out


def parse_args():
    p = argparse.ArgumentParser(
        description='给雪票财年 xlsx 的「年度雪票」sheet 末尾追加 4 列',
        formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__)
    p.add_argument('--xlsx', required=True, help='目标 xlsx 路径')
    p.add_argument('--shop', required=True, help='店铺名（DB order.shop，与 xlsx 内数据店铺一致）')
    p.add_argument('--start', default='2025-05-01', help='biz_date 起始（含），默认 2025-05-01')
    p.add_argument('--end', default='2026-04-30', help='biz_date 截止（含），默认 2026-04-30')
    return p.parse_args()


def main():
    args = parse_args()
    xlsx = os.path.abspath(args.xlsx)
    # 把 --end (含) 转半开区间 end_excl = end + 1 day
    from datetime import datetime, timedelta
    end_excl = (datetime.strptime(args.end, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')

    if not os.path.exists(xlsx):
        raise SystemExit(f'xlsx 不存在: {xlsx}')

    print(f'读 SQL ski_pass 聚合数据（{args.shop} / {args.start} ~ {end_excl}）...')
    by_code = fetch_skipass_by_order(args.shop, args.start, end_excl)
    print(f'  覆盖订单数: {len(by_code)}')

    print(f'打开 {xlsx}')
    wb = load_workbook(xlsx)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f'sheet {SHEET!r} 不在: {wb.sheetnames}')
    ws = wb[SHEET]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # 找「订单号」列定位 join 键
    if '订单号' not in headers:
        raise SystemExit('找不到「订单号」列')
    code_col = headers.index('订单号') + 1
    print(f'  「订单号」列在第 {code_col} 列')

    # 幂等：若末尾 4 列已是 NEW_COLS（任意顺序匹配），原地覆盖
    if len(headers) >= 4 and headers[-4:] == NEW_COLS:
        print('  检测到末尾已有目标 4 列，原地覆盖（幂等）')
        start_col = ws.max_column - 3
    else:
        start_col = ws.max_column + 1
        print(f'  追加 4 列到第 {start_col}~{start_col + 3} 列')

    # 取参考样式：第 1 行第 1 列（表头）
    ref_header = ws.cell(row=1, column=1)
    header_font = Font(bold=True, color='FFFFFF',
                       name=ref_header.font.name or 'Calibri',
                       size=ref_header.font.size or 11)
    header_fill = PatternFill('solid', fgColor=HEADER_COLOR)
    center = Alignment(horizontal='center', vertical='center')

    # 写表头
    for k, name in enumerate(NEW_COLS):
        cell = ws.cell(row=1, column=start_col + k, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    hits = misses = 0
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=code_col).value
        key = code.strip() if isinstance(code, str) else None
        rec = by_code.get(key) if key else None
        if rec:
            hits += 1
            name, deal, ticket, pick = rec
        else:
            misses += 1
            name = deal = ticket = pick = None

        ws.cell(row=r, column=start_col + 0, value=name)
        c_deal = ws.cell(row=r, column=start_col + 1, value=deal)
        c_tic = ws.cell(row=r, column=start_col + 2, value=ticket)
        c_pick = ws.cell(row=r, column=start_col + 3, value=pick)
        if deal is not None:
            c_deal.number_format = '0.00'
        if ticket is not None:
            c_tic.number_format = '0.00'
        if pick is not None:
            c_pick.number_format = 'yyyy-mm-dd hh:mm:ss'

    # 列宽自适应（粗略，按 max 字符宽度，上限 36）
    for k in range(4):
        col_letter = ws.cell(row=1, column=start_col + k).column_letter
        max_len = len(NEW_COLS[k]) * 2  # 中文字符按 2 宽
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=start_col + k).value
            if v is None:
                continue
            s = v.strftime('%Y-%m-%d %H:%M:%S') if hasattr(v, 'strftime') else str(v)
            w = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if w > max_len:
                max_len = w
        ws.column_dimensions[col_letter].width = min(max_len + 2, 36)

    print(f'  数据行命中订单数据: {hits}，未命中（无 ski_pass 行的订单留空）: {misses}')

    print(f'保存 {xlsx}')
    wb.save(xlsx)
    print('完成')


if __name__ == '__main__':
    main()
