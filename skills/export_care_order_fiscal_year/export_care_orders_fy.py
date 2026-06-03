#!/usr/bin/env python3
"""按店铺 + 财年口径导出年度养护（care）订单（单 sheet 宽表，财务/业务视角）。

仿 ../export_rent_order_fiscal_year（租赁财年版）规则，换成养护业务：
- 主表共用 [order]（o.type=N'养护'，万龙养护单 shop='万龙服务中心'）
- 行项目是 care（一张订单可含多块送修板）+ care_task（工序流水），无 rental_detail/
  charge_type/押金/超时费/赔偿金。养护毛收费 = repair_charge + common_charge。
- 支付/退款/分账/减免共用 order_payment/payment_refund/discount/order_share/payment_share
- 行粒度 = 一行一订单（多块养护在订单行聚合，不展开逐块明细）
- 营/非 照搬租赁雪季营业窗口（SEASON）；非雪季养护单 biz_date 落窗口外 → 非营业（设计预期）

单 sheet「年度养护」5 段拼接：固定前缀 + 动态支付区 + 动态退款区 + 固定中段(养护特化) + 固定后缀。
支付/退款明细列数 = 导出区间内单订单最大成功支付 / 有效退款笔数（数据驱动）。

用法：
    python export_care_orders_fy.py --shop 万龙服务中心
    python export_care_orders_fy.py --shop 万龙服务中心 --start 2025-05-01 --end 2026-04-30 --out x.xlsx

环境（Windows）：pip install pyodbc openpyxl；需 "ODBC Driver 18 for SQL Server"。
日期按 order.biz_date 过滤（财年报表语义）。数据源默认硬编码生产（复用
../export_rent_order/DEFAULT_CONN），可用 --conn 覆盖。
"""
import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta, date

import pyodbc
from openpyxl import Workbook

# 复用 ../export_rent_order 的单点真理（SHOP_PREFIX / REFUND_COND / DEFAULT_CONN / write_sheet）
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'export_rent_order'))
try:
    from export_rent_orders import SHOP_PREFIX, REFUND_COND, DEFAULT_CONN, write_sheet
except ImportError as e:
    print(f'无法 import ../export_rent_order/export_rent_orders.py：{e}\n'
          f'本 skill 依赖 sibling 目录 export_rent_order，请确认两个 skill 目录关系未被破坏。',
          file=sys.stderr)
    raise

sys.stdout.reconfigure(encoding='utf-8')

# ───────────────────────── 财年营业区间表（与租赁财年版一致，照搬雪季窗口） ─────────────────────────
# key = 财年标签 'YY-YY'（biz_date 落在 5/1~次年4/30）；value = (雪季营业起, 雪季营业止) inclusive
# 仅用于「营/非」「运营日序号」派生。非雪季养护单 biz_date 在窗口外 → 非营业（设计预期，用户已确认）。
SEASON = {
    '25-26': (date(2025, 10, 21), date(2026, 4, 9)),
}

SHEET_TITLE = '年度养护'
HEADER_COLOR = '1F4E78'

# ───────────────────────── SQL ─────────────────────────
# 通用过滤：单店 + 养护 + 业务日期区间 + code 非空（+ 可选 order.valid=1）。
# code 非空恒过滤：无 code 多是未下单/废弃单，不进财务报表（即便 --include-invalid 也排除）。
# __VALID__ 运行期占位（穿过 f-string，main 替换）：默认 'AND o.valid = 1'；--include-invalid → ''。
# valid 放宽仅作用于 order 表；care/order_payment/discount/order_share 等不变。
ORDER_FILTER = (
    "o.shop = ? AND o.[type] = N'养护' "
    "AND o.biz_date >= ? AND o.biz_date < ? "
    "AND o.code IS NOT NULL AND LTRIM(RTRIM(o.code)) <> N'' __VALID__"
)

# 主查询：一行一订单。订单级列全部用聚合/标量子查询保订单粒度。
MAIN_SQL = f"""
SELECT
    o.id                                                   AS _oid,
    o.valid                                                AS _valid,
    o.code                                                 AS 订单号,
    o.biz_date                                             AS _biz_dt,
    o.create_date                                          AS _crt_dt,
    ROW_NUMBER() OVER (PARTITION BY o.shop, CAST(o.biz_date AS DATE)
                       ORDER BY o.create_date ASC, o.id ASC) AS 日序号,
    ISNULL(pay_agg.cnt, 0)                                  AS 支付次数,
    ISNULL(pay_agg.paid, 0)                                 AS 支付合计,
    ISNULL(ref_agg.cnt, 0)                                  AS 退款次数,
    ISNULL(ref_agg.refund, 0)                               AS 退款合计,
    ISNULL(pay_agg.paid, 0) - ISNULL(ref_agg.refund, 0)     AS 订单结余,
    ref_agg.last_refund_dt                                  AS _last_refund_dt,
    ISNULL(ca.care_cnt, 0)                                  AS 养护件数,
    ISNULL(ca.repair_sum, 0)                                AS 维修费合计,
    ISNULL(ca.common_sum, 0)                                AS 普通养护费合计,
    ISNULL((
        SELECT SUM(dd.amount) FROM (
            SELECT DISTINCT d.id, d.amount
            FROM discount d
            WHERE d.valid = 1 AND (
                  d.order_id = o.id
               OR (d.biz_type = N'养护' AND d.biz_id IN (
                     SELECT c.id FROM care c WHERE c.order_id = o.id AND c.valid = 1))
            )
        ) dd
    ), 0)                                                   AS 减免合计,
    ISNULL(ca.ticket_disc_sum, 0)                           AS 卡券减免合计,
    ISNULL(ca.care_disc_sum, 0)                             AS 养护直减合计,
    ISNULL(ca.f_edge, 0)                                    AS _f_edge,
    ISNULL(ca.f_wax, 0)                                     AS _f_wax,
    ISNULL(ca.f_unwax, 0)                                   AS _f_unwax,
    ISNULL(ca.f_repair, 0)                                  AS _f_repair,
    CASE WHEN o.hide = 1 THEN N'是' ELSE N'' END            AS 隐藏订单,
    ISNULL(osh.should_share, 0)                             AS 应分账金额,
    ISNULL(psh.actual_share, 0)                             AS 实分账金额,
    ISNULL(osh.should_share, 0) - ISNULL(psh.actual_share, 0) AS 待分账金额,
    o.shop                                                  AS 门店,
    COALESCE(NULLIF(LTRIM(RTRIM(o.contact_name)), N''), m.real_name) AS 客户名称,
    COALESCE(NULLIF(LTRIM(RTRIM(o.contact_num)), N''), msa_cell.num) AS 电话,
    msa_oid.num                                             AS 顾客openid,
    big_pay.pay_method                                      AS 收款方式,
    CASE WHEN big_pay.pay_method = N'微信支付' THEN wk.mch_id ELSE NULL END AS 支付账号,
    CAST(o.biz_date AS DATE)                                AS 业务日期,
    CONVERT(VARCHAR(8), o.biz_date, 108)                    AS 业务时间,
    s.name                                                  AS 店员姓名,
    ISNULL(NULLIF(LTRIM(RTRIM(staff_oid.openid)), N''), N'') AS 店员openid,
    CASE WHEN ISNULL(pay_agg.paid, 0) < 5
              OR (s.name IS NOT NULL AND s.name LIKE N'%苍%')
         THEN N'是' ELSE N'' END                            AS 测试,
    ISNULL(cs.care_cnt, 0)                                  AS _care_cnt,
    ISNULL(cs.done_cnt, 0)                                  AS _done_cnt,
    ISNULL(cs.notstart_cnt, 0)                              AS _ns_cnt,
    CASE WHEN o.entertain = 1 OR EXISTS (
             SELECT 1 FROM care c WHERE c.order_id = o.id
                                    AND c.valid = 1 AND c.entertain = 1)
         THEN 1 ELSE 0 END                                  AS _entertain
FROM [order] o
LEFT JOIN staff s  ON s.id = o.staff_id
LEFT JOIN member m ON m.id = o.member_id
LEFT JOIN (
    SELECT order_id, COUNT(*) AS cnt, SUM(amount) AS paid
    FROM order_payment WHERE status = N'支付成功' AND valid = 1
    GROUP BY order_id
) pay_agg ON pay_agg.order_id = o.id
LEFT JOIN (
    SELECT pr.order_id, COUNT(*) AS cnt, SUM(pr.amount) AS refund,
           MAX(pr.create_date) AS last_refund_dt
    FROM payment_refund pr WHERE {REFUND_COND}
    GROUP BY pr.order_id
) ref_agg ON ref_agg.order_id = o.id
LEFT JOIN (
    SELECT c.order_id,
           COUNT(*) AS care_cnt,
           ISNULL(SUM(c.repair_charge), 0)   AS repair_sum,
           ISNULL(SUM(c.common_charge), 0)   AS common_sum,
           ISNULL(SUM(c.discount), 0)        AS care_disc_sum,
           ISNULL(SUM(c.ticket_discount), 0) AS ticket_disc_sum,
           MAX(CASE WHEN c.need_edge = 1 THEN 1 ELSE 0 END)   AS f_edge,
           MAX(CASE WHEN c.need_wax = 1 THEN 1 ELSE 0 END)    AS f_wax,
           MAX(CASE WHEN c.need_unwax = 1 THEN 1 ELSE 0 END)  AS f_unwax,
           MAX(CASE WHEN c.need_repair = 1 THEN 1 ELSE 0 END) AS f_repair
    FROM care c WHERE c.valid = 1
    GROUP BY c.order_id
) ca ON ca.order_id = o.id
LEFT JOIN (
    SELECT order_id, SUM(amount) AS should_share
    FROM order_share WHERE valid = 1 GROUP BY order_id
) osh ON osh.order_id = o.id
LEFT JOIN (
    SELECT os2.order_id, SUM(ps.amount) AS actual_share
    FROM payment_share ps JOIN order_share os2 ON os2.id = ps.share_id
    WHERE ps.valid = 1 AND ps.success = 1
    GROUP BY os2.order_id
) psh ON psh.order_id = o.id
LEFT JOIN (
    -- 订单级养护状态聚合：每块 care 取最后一条 care_task（valid=1，按 id DESC）= Care.cs currentStep；
    -- task_name ∈ (发板,强行索回) → 该 care 已完成。done/notstart 计数供 derive_care_status 派生。
    SELECT cc.order_id,
           COUNT(*) AS care_cnt,
           SUM(CASE WHEN lt.task_name IS NULL THEN 0
                    WHEN LTRIM(RTRIM(lt.task_name)) IN (N'发板', N'强行索回') THEN 1
                    ELSE 0 END) AS done_cnt,
           SUM(CASE WHEN lt.task_name IS NULL THEN 1 ELSE 0 END) AS notstart_cnt
    FROM care cc
    OUTER APPLY (
        SELECT TOP 1 ct.task_name FROM care_task ct
        WHERE ct.care_id = cc.id AND ct.valid = 1
        ORDER BY ct.id DESC
    ) lt
    WHERE cc.valid = 1
    GROUP BY cc.order_id
) cs ON cs.order_id = o.id
OUTER APPLY (
    SELECT TOP 1 op.pay_method, op.mch_id
    FROM order_payment op
    WHERE op.order_id = o.id AND op.status = N'支付成功' AND op.valid = 1
    ORDER BY op.amount DESC, op.id ASC
) big_pay
LEFT JOIN wepay_key wk ON wk.id = big_pay.mch_id
OUTER APPLY (
    SELECT TOP 1 num FROM member_social_account
    WHERE member_id = o.member_id AND type = N'cell' AND valid = 1
      AND LTRIM(RTRIM(num)) <> N'' ORDER BY id ASC
) msa_cell
OUTER APPLY (
    SELECT TOP 1 num FROM member_social_account
    WHERE member_id = o.member_id AND type = N'wechat_mini_openid' AND valid = 1
      AND LTRIM(RTRIM(num)) <> N'' ORDER BY id ASC
) msa_oid
OUTER APPLY (
    SELECT TOP 1 saj.wechat_mini_openid AS openid
    FROM staff_social_account ssa
    JOIN social_account_for_job saj ON saj.id = ssa.social_account_id
    -- 历史归集口径：不过滤 ssa.valid（离职店员旧账号 valid=0，历史经手订单仍要还原）。
    -- 两级偏好：① 优先取日期窗口覆盖该 biz_date 的账号；② 若无任何覆盖窗口，
    -- 回退到该店员「曾用过」的最近账号（start_date DESC）。仅排除 openid 空白记录。
    WHERE ssa.staff_id = o.staff_id
      AND LTRIM(RTRIM(saj.wechat_mini_openid)) <> N''
    ORDER BY
      CASE WHEN CAST(ssa.start_date AS DATE) <= CAST(o.biz_date AS DATE)
                AND (ssa.end_date IS NULL OR CAST(ssa.end_date AS DATE) >= CAST(o.biz_date AS DATE))
           THEN 0 ELSE 1 END,
      ssa.start_date DESC, ssa.id DESC
) staff_oid
WHERE {ORDER_FILTER}
ORDER BY o.biz_date ASC, o.create_date ASC, o.id ASC
"""

# 支付明细（段2）：每笔成功有效支付，按支付时间升序
PAY_DETAIL_SQL = f"""
SELECT o.id AS oid,
       COALESCE(op.paid_date, op.create_date) AS pt,
       op.amount,
       op.pay_method,
       CASE WHEN op.pay_method = N'微信支付' THEN wk.mch_id ELSE NULL END AS acct
FROM [order] o
JOIN order_payment op ON op.order_id = o.id
LEFT JOIN wepay_key wk ON wk.id = op.mch_id
WHERE {ORDER_FILTER} AND op.status = N'支付成功' AND op.valid = 1
ORDER BY o.id ASC, COALESCE(op.paid_date, op.create_date) ASC, op.id ASC
"""

# 退款明细（段3）：每笔有效退款，退款方式取原支付通道 pay_method（payment_refund 无方式列）
# 退款账号：微信支付→wepay_key.mch_id；支付宝→"支付宝"；其他→空
# 退款人：pr.staff_id → staff.name
REFUND_DETAIL_SQL = f"""
SELECT o.id AS oid,
       pr.create_date AS rt,
       pr.amount,
       op.pay_method AS refund_method,
       CASE WHEN op.pay_method = N'微信支付' THEN wk.mch_id
            WHEN op.pay_method = N'支付宝'   THEN N'支付宝'
            ELSE N'' END                            AS refund_account,
       sa.name                                       AS refund_staff
FROM [order] o
JOIN payment_refund pr ON pr.order_id = o.id
LEFT JOIN order_payment op ON op.id = pr.payment_id
LEFT JOIN wepay_key wk ON wk.id = op.mch_id
LEFT JOIN staff sa ON sa.id = pr.staff_id
WHERE {ORDER_FILTER} AND {REFUND_COND}
ORDER BY o.id ASC, pr.create_date ASC, pr.id ASC
"""

# 主查询里以 _ 前缀的辅助列（不进 Excel，仅供 Python 派生）
AUX_COLS = {'_oid', '_valid', '_biz_dt', '_crt_dt', '_last_refund_dt',
            '_f_edge', '_f_wax', '_f_unwax', '_f_repair',
            '_care_cnt', '_done_cnt', '_ns_cnt', '_entertain'}

# 需要 round(2) + number_format='0.00' 的金额列名（固定列部分；动态列另判）
MONEY_FIXED = {'支付合计', '退款合计', '订单结余', '维修费合计', '普通养护费合计',
               '减免合计', '卡券减免合计', '养护直减合计',
               '应分账金额', '实分账金额', '待分账金额'}

# need_* 标志 → 服务项目标签
SERVICE_LABELS = [('_f_edge', '修刃'), ('_f_wax', '打蜡'),
                  ('_f_unwax', '去蜡'), ('_f_repair', '维修')]


def fiscal_year(d):
    """biz_date → 财年标签 'YY-YY'（5/1 起，次年 4/30 止）"""
    y = d.year
    fy = y if d.month >= 5 else y - 1
    return f'{fy % 100:02d}-{(fy + 1) % 100:02d}'


def split_dt(v):
    """datetime → (date, 'HH:MM:SS')；None → (None, None)"""
    if v is None:
        return None, None
    return v.date(), v.strftime('%H:%M:%S')


def derive_care_status(care_cnt, done_cnt, ns_cnt):
    """订单级养护状态近似（复刻 Care.cs 计算属性 status 并聚合到订单级）。
    单 care 状态：最后一条 care_task(valid=1, id DESC) task_name ∈ (发板,强行索回) → 已完成；
    无 care_task → 未开始；否则 进行中。订单级：
      - 无有效 care → ''（后端 careProperties 为 null）
      - 全部已完成 → 已完成
      - 全部未开始 → 未开始
      - 0 已完成但有进行中 → 进行中
      - 部分已完成 → 部分完成
    ⚠️ 近似值（care.finish 生产恒 0，状态真信号在 care_task 工序），使用前抽样验收。"""
    if care_cnt == 0:
        return ''
    if done_cnt == care_cnt:
        return '已完成'
    if done_cnt == 0 and ns_cnt == care_cnt:
        return '未开始'
    if done_cnt == 0:
        return '进行中'
    return '部分完成'


def parse_args():
    p = argparse.ArgumentParser(
        description='按店铺+财年导出年度养护订单（单 sheet 宽表）',
        formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__)
    p.add_argument('--shop', required=True, help='店铺名（DB order.shop，如「万龙服务中心」）')
    p.add_argument('--start', default='2025-05-01', help='业务起始日期 biz_date inclusive，默认 2025-05-01')
    p.add_argument('--end', default='2026-04-30', help='业务截止日期 biz_date inclusive，默认 2026-04-30')
    p.add_argument('--out', default=None, help='输出 xlsx 路径，默认 {prefix}_care_orders_fy_{start}_{end}.xlsx')
    p.add_argument('--conn', default=DEFAULT_CONN, help='ODBC 连接串，默认连生产')
    p.add_argument('--include-invalid', action='store_true',
                   help='导出 order 不论 valid 是否=1（默认仅 valid=1）。'
                        '仅放宽 order 表；care/支付/退款/分账等 valid 过滤不变')
    return p.parse_args()


def default_out_name(shop, start, end):
    prefix = SHOP_PREFIX.get(shop, shop)
    return f'{prefix}_care_orders_fy_{start}_{end}.xlsx'


def main():
    args = parse_args()
    start = args.start
    end_excl = (datetime.strptime(args.end, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
    out = os.path.abspath(args.out or default_out_name(args.shop, args.start, args.end))
    params = (args.shop, start, end_excl)

    valid_clause = '' if args.include_invalid else 'AND o.valid = 1'
    main_sql = MAIN_SQL.replace('__VALID__', valid_clause)
    pay_detail_sql = PAY_DETAIL_SQL.replace('__VALID__', valid_clause)
    refund_detail_sql = REFUND_DETAIL_SQL.replace('__VALID__', valid_clause)
    print(f'order.valid 过滤: {"关闭（不论 valid 全导）" if args.include_invalid else "valid=1"}')

    print('连数据库 ...')
    cn = pyodbc.connect(args.conn)
    cur = cn.cursor()

    print(f'主查询（{args.shop} / 业务 {start} ~ {end_excl}） ...')
    cur.execute(main_sql, *params)
    main_cols = [c[0] for c in cur.description]
    main_rows = cur.fetchall()
    print(f'  订单: {len(main_rows)} 行')

    print('支付明细 ...')
    cur.execute(pay_detail_sql, *params)
    pay_by_oid = defaultdict(list)
    for oid, pt, amt, pm, acct in cur.fetchall():
        pay_by_oid[oid].append((pt, amt, pm, acct))
    print(f'  支付笔: {sum(len(v) for v in pay_by_oid.values())}')

    print('退款明细 ...')
    cur.execute(refund_detail_sql, *params)
    ref_by_oid = defaultdict(list)
    for oid, rt, amt, rm, racct, rstaff in cur.fetchall():
        ref_by_oid[oid].append((rt, amt, rm, racct, rstaff))
    print(f'  退款笔: {sum(len(v) for v in ref_by_oid.values())}')
    cn.close()

    idx = {name: i for i, name in enumerate(main_cols)}
    I_OID, I_VALID = idx['_oid'], idx['_valid']
    I_CODE, I_PAYCNT = idx['订单号'], idx['支付次数']

    # ── 同订单号去重（用户指定优先级）──
    # 1) 有成功支付记录的那条；2) 否则 valid=1 的那条；3) 否则 id 最大的那条
    groups = defaultdict(list)
    for r in main_rows:
        groups[str(r[I_CODE]).strip()].append(r)
    deduped, dropped, dup_codes = [], 0, 0
    for code, grp in groups.items():
        if len(grp) > 1:
            dup_codes += 1
            dropped += len(grp) - 1
            winner = max(grp, key=lambda r: (
                1 if (r[I_PAYCNT] or 0) > 0 else 0,
                1 if r[I_VALID] == 1 else 0,
                r[I_OID]))
            deduped.append(winner)
        else:
            deduped.append(grp[0])
    deduped.sort(key=lambda r: (r[idx['_biz_dt']], r[idx['_crt_dt']], r[I_OID]))
    print(f'  重复订单号 {dup_codes} 个，去重丢弃 {dropped} 行 → 保留 {len(deduped)} 行')
    main_rows = deduped

    kept_oids = [r[I_OID] for r in main_rows]
    max_pay = max((len(pay_by_oid.get(o, [])) for o in kept_oids), default=0)
    max_refund = max((len(ref_by_oid.get(o, [])) for o in kept_oids), default=0)
    print(f'  maxPay={max_pay}  maxRefund={max_refund}（按去重后集）')

    # 财年缺失校验：所有 biz_date 的财年都要在 SEASON
    missing = {}
    for row in main_rows:
        bd = row[idx['_biz_dt']]
        if bd is None:
            continue
        fy = fiscal_year(bd)
        if fy not in SEASON:
            missing[fy] = missing.get(fy, 0) + 1
    if missing:
        det = '，'.join(f'{k}（{v} 单）' for k, v in sorted(missing.items()))
        raise SystemExit(
            f'以下财年不在 SEASON 营业区间表中：{det}。\n'
            f'请在 export_care_orders_fy.py 的 SEASON dict 补对应「财年→(营业起,营业止)」后重跑。')

    # ── 组装表头（与数据行同一处生成，杜绝错位）──
    headers = ['业务', '财年', '营/非', '财年序号', '运营日序号', '日序号', '月份',
               '创建日期', '创建时间', '支付次数', '支付合计', '退款次数', '退款合计',
               '订单结余', '订单状态', '最后退款日期', '最后退款时间']
    for k in range(1, max_pay + 1):
        headers += [f'【支付{k}】日期', f'【支付{k}】时间', f'【支付{k}】金额',
                    f'【支付{k}】支付方式', f'【支付{k}】支付账号']
    for k in range(1, max_refund + 1):
        headers += [f'【退款{k}】日期', f'【退款{k}】时间', f'【退款{k}】金额',
                    f'【退款{k}】退款方式',
                    f'【退款{k}】退款账号', f'【退款{k}】退款人']
    # 段4（养护特化：养护件数/三类费用/服务项目 + 分账/会员/收款，照搬租赁段4其余列）
    headers += ['养护件数', '维修费合计', '普通养护费合计', '减免合计',
                '卡券减免合计', '养护直减合计', '服务项目', '隐藏订单',
                '应分账金额', '实分账金额', '待分账金额', '业务', '门店',
                '客户名称', '电话', '顾客openid', '收款方式', '支付账号']
    # 段5（照搬租赁后缀 14 列）
    headers += ['订单号', '业务日期', '业务时间', '结算日期', '结算时间',
                '支付总金额', '退款总金额', '订单结余', '店员姓名', '店员openid', '测试',
                '临时订单', '客户名称', '正/闭']

    money_cols = set()
    for ci, h in enumerate(headers, start=1):
        if h in MONEY_FIXED or h in ('支付总金额', '退款总金额') \
           or h.endswith('】金额') or (h == '订单结余'):
            money_cols.add(ci)

    rows = []
    for row in main_rows:
        g = lambda n: row[idx[n]]
        oid = g('_oid')
        bd = g('_biz_dt')
        fy = fiscal_year(bd) if bd else ''
        s_start, s_end = (SEASON.get(fy, (None, None)))
        bdate = bd.date() if bd else None
        if bdate and s_start and s_start <= bdate <= s_end:
            ying = '营业'
            ops_day = (bdate - s_start).days + 1
        else:
            ying = '非营业'
            ops_day = None
        crt_d, crt_t = split_dt(g('_crt_dt'))
        lr_d, lr_t = split_dt(g('_last_refund_dt'))
        paid = g('支付合计') or 0
        refund = g('退款合计') or 0
        balance = g('订单结余') or 0
        test_flag = g('测试')
        care_cnt = g('_care_cnt') or 0
        status = derive_care_status(care_cnt, g('_done_cnt') or 0, g('_ns_cnt') or 0)
        temp_order = '是' if (test_flag != '是' and balance > 0 and care_cnt == 0) else ''
        zheng_bi = '关闭' if (paid == 0 and (g('_entertain') or 0) == 0) else '正常'
        service = '·'.join(lbl for aux, lbl in SERVICE_LABELS if (g(aux) or 0) == 1)

        seg1 = ['养护', fy, ying, '', ops_day, g('日序号'), (bd.month if bd else None),
                crt_d, crt_t, g('支付次数') or 0, round(float(paid), 2),
                g('退款次数') or 0, round(float(refund), 2), round(float(balance), 2),
                status, lr_d, lr_t]

        seg2 = []
        plist = pay_by_oid.get(oid, [])
        for k in range(max_pay):
            if k < len(plist):
                pt, amt, pm, acct = plist[k]
                pd_, ptm = split_dt(pt)
                seg2 += [pd_, ptm, round(float(amt), 2) if amt is not None else None, pm, acct]
            else:
                seg2 += [None, None, None, None, None]

        seg3 = []
        rlist = ref_by_oid.get(oid, [])
        for k in range(max_refund):
            if k < len(rlist):
                rt, amt, rm, racct, rstaff = rlist[k]
                rd_, rtm = split_dt(rt)
                seg3 += [rd_, rtm, round(float(amt), 2) if amt is not None else None,
                         rm, racct or '', rstaff or '']
            else:
                seg3 += [None, None, None, None, None, None]

        seg4 = [care_cnt, round(float(g('维修费合计') or 0), 2),
                round(float(g('普通养护费合计') or 0), 2), round(float(g('减免合计') or 0), 2),
                round(float(g('卡券减免合计') or 0), 2), round(float(g('养护直减合计') or 0), 2),
                service, g('隐藏订单'),
                round(float(g('应分账金额') or 0), 2), round(float(g('实分账金额') or 0), 2),
                round(float(g('待分账金额') or 0), 2), '养护', g('门店'),
                g('客户名称'), g('电话'), g('顾客openid'), g('收款方式'), g('支付账号')]

        seg5 = [g('订单号'), g('业务日期'), g('业务时间'), lr_d, lr_t,
                round(float(paid), 2), round(float(refund), 2), round(float(balance), 2),
                g('店员姓名'), g('店员openid'), test_flag, temp_order, g('客户名称'), zheng_bi]

        full = seg1 + seg2 + seg3 + seg4 + seg5
        assert len(full) == len(headers), f'列错位 row={len(full)} header={len(headers)}'
        rows.append(full)

    print(f'写 Excel: {out}  （{len(headers)} 列 × {len(rows)} 行）')
    wb = Workbook()
    write_sheet(wb.active, SHEET_TITLE, HEADER_COLOR, headers, rows)
    ws = wb.active
    for ci in money_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=ci).number_format = '0.00'
    wb.save(out)
    print(f'  完成。文件大小: {os.path.getsize(out) / 1024:.1f} KB')
    print(f'完成: {out}')


if __name__ == '__main__':
    main()
