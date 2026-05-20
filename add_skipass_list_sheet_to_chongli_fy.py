"""把 雪票列表_2026-05-20.xls 作为第 4 个 sheet 加入
chongli_ski_pass_orders_fy_2025-05-01_2026-04-30.xlsx。

- 源 xls：xlrd 读，1 sheet × 28 列 × 614 行（含表头）
- 目标 xlsx：openpyxl 写，新 sheet 名「雪票列表」
- 表头沿用蓝底白字粗体（1F4E78）+ 居中
- 数据原样拷贝（日期是字符串就保字符串，数值保数值）
- 幂等：目标已有同名 sheet 则先删再建
- 其他 3 sheet（年度雪票 / 支付明细 / 支付流水）不动
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

XLS_SRC = r'D:\snowmeet\snowmeet_ai_doc\雪票列表_2026-05-20.xls'
XLSX_DST = r'D:\snowmeet\snowmeet_ai_doc\chongli_ski_pass_orders_fy_2025-05-01_2026-04-30.xlsx'
NEW_SHEET = '雪票列表'
HEADER_COLOR = '1F4E78'


def visual_width(s):
    return sum(2 if ord(ch) > 127 else 1 for ch in str(s))


def main():
    if not os.path.exists(XLS_SRC):
        raise SystemExit(f'源 xls 不存在: {XLS_SRC}')
    if not os.path.exists(XLSX_DST):
        raise SystemExit(f'目标 xlsx 不存在: {XLSX_DST}')

    print(f'读 {XLS_SRC}')
    src_wb = xlrd.open_workbook(XLS_SRC)
    src_ws = src_wb.sheet_by_index(0)
    nrows, ncols = src_ws.nrows, src_ws.ncols
    print(f'  源 sheet「{src_wb.sheet_names()[0]}」: {nrows} 行 × {ncols} 列')

    print(f'打开 {XLSX_DST}')
    dst_wb = load_workbook(XLSX_DST)
    print(f'  原 sheets: {dst_wb.sheetnames}')

    if NEW_SHEET in dst_wb.sheetnames:
        print(f'  已存在「{NEW_SHEET}」sheet，删除重建（幂等）')
        del dst_wb[NEW_SHEET]

    ws = dst_wb.create_sheet(NEW_SHEET)

    header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    header_fill = PatternFill('solid', fgColor=HEADER_COLOR)
    center = Alignment(horizontal='center', vertical='center')

    # 表头（第 1 行）
    headers = [src_ws.cell_value(0, c) for c in range(ncols)]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 数据行
    for r in range(1, nrows):
        for c in range(ncols):
            v = src_ws.cell_value(r, c)
            # xlrd 的 number type 都是 float；空串 -> None 更干净
            if isinstance(v, str) and v == '':
                v = None
            ws.cell(row=r + 1, column=c + 1, value=v)

    # 列宽自适应（粗略，上限 36）
    for c in range(1, ncols + 1):
        max_w = visual_width(headers[c - 1])
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            w = visual_width(v)
            if w > max_w:
                max_w = w
        ws.column_dimensions[get_column_letter(c)].width = min(max_w + 2, 36)

    # 冻结表头
    ws.freeze_panes = 'A2'

    print(f'  新 sheet 行数 = {ws.max_row}, 列数 = {ws.max_column}')

    print(f'保存 {XLSX_DST}')
    dst_wb.save(XLSX_DST)
    print(f'  保存后 sheets: {load_workbook(XLSX_DST, read_only=True).sheetnames}')
    print('完成')


if __name__ == '__main__':
    main()
