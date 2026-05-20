"""给 {prefix}_ski_pass_orders_fy_{start}_{end}.xlsx 添加第 4 个 sheet「年度雪票明细」：
  关联「年度雪票」（订单级） + ski_pass 表（雪票级明细）。

明细列（9）：
  雪票名称 (product_name) / 票价 (ticket_price) / 押金 (deposit) /
  退款金额 (refund_amount) / 是否退款 (have_refund: 1→是 / NULL→否) /
  取卡日期 + 取卡时间（拆 card_member_pick_time）/
  退卡日期 + 退卡时间（拆 card_member_return_time）

形态：
  - 订单级列（即「年度雪票」原全部 N 列）放左侧，一单多票纵向合并单元格
  - 明细 9 列在右侧，每张票各一行
  - 一单无 ski_pass 行（如南山 389 单空订单）→ 保留单行，明细列空
  - 多明细订单（M ≥ 2）整行底色浅蓝 EAF2FB（视觉提示）
  - 表头蓝底白字粗体 1F4E78；freeze A2

幂等：「年度雪票明细」存在则删重建；其他 3 sheet（年度雪票/支付明细/支付流水）不动。

用法：
  py add_skipass_detail_merged_sheet.py --xlsx nanshan_ski_pass_orders_fy_2025-05-01_2026-04-30.xlsx --shop 南山
  py add_skipass_detail_merged_sheet.py --xlsx chongli_ski_pass_orders_fy_2025-05-01_2026-04-30.xlsx --shop 崇礼旗舰店
"""
import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta

import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

MAIN_SHEET = '年度雪票'
DETAIL_SHEET = '年度雪票明细'
HEADER_COLOR = '1F4E78'
MULTI_FILL = 'EAF2FB'   # 多明细订单底色（与零售明细 sibling 一致）

DETAIL_COLS = ['雪票名称', '票价', '押金', '退款金额', '是否退款',
               '取卡日期', '取卡时间', '退卡日期', '退卡时间']

MONEY_DETAIL = {'票价', '押金', '退款金额'}

CONN = ('DRIVER={ODBC Driver 18 for SQL Server};SERVER=tcp:100.28.143.19,1433;'
        'DATABASE=snowmeet_new;UID=claude;PWD=abcd123!@#;'
        'Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=30;')


def split_dt(v):
    """返回 (date, time_str) 或 (None, None)"""
    if v is None:
        return None, None
    return v.date(), v.strftime('%H:%M:%S')


def parse_args():
    p = argparse.ArgumentParser(
        description='给雪票财年 xlsx 添加「年度雪票明细」合并 sheet',
        formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__)
    p.add_argument('--xlsx', required=True, help='目标 xlsx 路径')
    p.add_argument('--shop', required=True, help='店铺名（DB order.shop）')
    p.add_argument('--start', default='2025-05-01', help='biz_date 起始（含），默认 2025-05-01')
    p.add_argument('--end', default='2026-04-30', help='biz_date 截止（含），默认 2026-04-30')
    return p.parse_args()


def fetch_skipass_rows(shop, start, end_excl):
    """返回 {order_code: [(name, ticket_price, deposit, refund_amount, have_refund, pick_dt, return_dt), ...]}"""
    cn = pyodbc.connect(CONN)
    cur = cn.cursor()
    cur.execute("""SELECT o.code, sp.product_name, sp.ticket_price, sp.deposit,
                          sp.refund_amount, sp.have_refund,
                          sp.card_member_pick_time, sp.card_member_return_time,
                          sp.id
                   FROM ski_pass sp JOIN [order] o ON o.id = sp.order_id
                   WHERE o.shop = ? AND o.[type] = N'雪票'
                     AND o.biz_date >= ? AND o.biz_date < ?
                     AND o.valid = 1 AND o.code IS NOT NULL AND LTRIM(RTRIM(o.code)) <> N''
                     AND sp.valid = 1
                   ORDER BY o.id, sp.id""", shop, start, end_excl)
    by_code = defaultdict(list)
    for r in cur.fetchall():
        code = r[0].strip() if r[0] else None
        if not code:
            continue
        # (name, ticket, deposit, refund, have_refund, pick, return)
        by_code[code].append((r[1], r[2], r[3], r[4], r[5], r[6], r[7]))
    cn.close()
    return by_code


def main():
    args = parse_args()
    xlsx = os.path.abspath(args.xlsx)
    end_excl = (datetime.strptime(args.end, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')

    if not os.path.exists(xlsx):
        raise SystemExit(f'xlsx 不存在: {xlsx}')

    print(f'读 SQL ski_pass 明细（{args.shop} / {args.start} ~ {end_excl}）...')
    by_code = fetch_skipass_rows(args.shop, args.start, end_excl)
    total_rows = sum(len(v) for v in by_code.values())
    multi_orders = sum(1 for v in by_code.values() if len(v) > 1)
    print(f'  覆盖订单: {len(by_code)}（含 {multi_orders} 单多票），明细总行: {total_rows}')

    print(f'打开 {xlsx}')
    wb = load_workbook(xlsx)
    if MAIN_SHEET not in wb.sheetnames:
        raise SystemExit(f'缺主 sheet「{MAIN_SHEET}」: {wb.sheetnames}')

    # 幂等：删旧 DETAIL_SHEET
    if DETAIL_SHEET in wb.sheetnames:
        print(f'  「{DETAIL_SHEET}」已存在，删除重建（幂等）')
        del wb[DETAIL_SHEET]

    main_ws = wb[MAIN_SHEET]
    main_headers = [main_ws.cell(row=1, column=c).value for c in range(1, main_ws.max_column + 1)]
    n_order_cols = len(main_headers)
    print(f'  年度雪票订单级列数 N = {n_order_cols}')

    # 读主表所有数据行
    main_rows = []
    for r in range(2, main_ws.max_row + 1):
        row = [main_ws.cell(row=r, column=c).value for c in range(1, n_order_cols + 1)]
        main_rows.append(row)

    # 找主表「订单号」列
    if '订单号' not in main_headers:
        raise SystemExit('年度雪票找不到「订单号」列')
    code_idx = main_headers.index('订单号')   # 0-based

    # 找各金额列（订单级，用于格式化）
    money_order_cols = set()
    for ci, h in enumerate(main_headers, start=1):
        if h in ('支付合计', '退款合计', '订单结余', '成交价合计', '减免合计',
                 '应分账金额', '实分账金额', '待分账金额', '支付总金额', '退款总金额',
                 '支付价格', '结算价格') or (isinstance(h, str) and h.endswith('】金额')):
            money_order_cols.add(ci)

    # 找主表里的取票时间列（datetime 型，需保格式）
    datetime_order_cols = set()
    for ci, h in enumerate(main_headers, start=1):
        if h == '取票时间':
            datetime_order_cols.add(ci)

    # 新 sheet
    ws = wb.create_sheet(DETAIL_SHEET)
    header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    header_fill = PatternFill('solid', fgColor=HEADER_COLOR)
    multi_fill = PatternFill('solid', fgColor=MULTI_FILL)
    center = Alignment(horizontal='center', vertical='center')

    # 写表头：N 订单级 + 9 明细
    all_headers = list(main_headers) + DETAIL_COLS
    for ci, h in enumerate(all_headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    # 展开数据
    out_row = 2
    merges = []   # list of (col, start_row, end_row) — 订单级列需合并
    multi_row_ranges = []   # list of (start_row, end_row) — 多明细订单的行区间用于上色

    for main_row in main_rows:
        code = main_row[code_idx]
        key = code.strip() if isinstance(code, str) else None
        detail_list = by_code.get(key, []) if key else []
        M = max(len(detail_list), 1)   # 至少 1 行

        # 写订单级列（仅第一行写值，其余空 — 等同合并视觉）
        for ci, v in enumerate(main_row, start=1):
            ws.cell(row=out_row, column=ci, value=v)
        # 写明细列
        for k in range(M):
            r = out_row + k
            if detail_list and k < len(detail_list):
                name, tic, dep, ref, hrefund, pick, retn = detail_list[k]
                hrefund_label = '是' if hrefund == 1 else '否'
                pick_d, pick_t = split_dt(pick)
                ret_d, ret_t = split_dt(retn)
                details = [name, tic, dep, ref, hrefund_label, pick_d, pick_t, ret_d, ret_t]
            else:
                details = [None] * 9   # 该订单无 ski_pass 行 → 明细列空
            for di, dv in enumerate(details):
                c = ws.cell(row=r, column=n_order_cols + 1 + di, value=dv)
                # 金额格式
                col_name = DETAIL_COLS[di]
                if col_name in MONEY_DETAIL and dv is not None:
                    c.number_format = '0.00'

        # 多行订单：标记合并 + 上色
        if M > 1:
            for ci in range(1, n_order_cols + 1):
                merges.append((ci, out_row, out_row + M - 1))
            multi_row_ranges.append((out_row, out_row + M - 1))

        out_row += M

    last_row = out_row - 1
    print(f'  写入数据行: {last_row - 1} 行（含合并展开）')

    # 应用合并
    for ci, start_r, end_r in merges:
        col_letter = get_column_letter(ci)
        ws.merge_cells(f'{col_letter}{start_r}:{col_letter}{end_r}')

    # 多明细订单底色（整行上色，含明细列；同一订单 N 行视觉一体）
    for start_r, end_r in multi_row_ranges:
        for r in range(start_r, end_r + 1):
            for ci in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=ci)
                if cell.fill.fill_type is None:
                    cell.fill = multi_fill

    # 订单级金额列保留格式
    for ci in money_order_cols:
        for r in range(2, last_row + 1):
            cell = ws.cell(row=r, column=ci)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'

    # 订单级取票时间列保留 datetime 格式
    for ci in datetime_order_cols:
        for r in range(2, last_row + 1):
            cell = ws.cell(row=r, column=ci)
            if cell.value is not None and hasattr(cell.value, 'strftime'):
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    # freeze + 列宽
    ws.freeze_panes = 'A2'
    for ci in range(1, len(all_headers) + 1):
        col_letter = get_column_letter(ci)
        max_w = sum(2 if ord(ch) > 127 else 1 for ch in str(all_headers[ci - 1]))
        # 采样行计算列宽
        for r in range(2, min(last_row + 1, 200)):
            v = ws.cell(row=r, column=ci).value
            if v is None:
                continue
            s = v.strftime('%Y-%m-%d %H:%M:%S') if hasattr(v, 'strftime') else str(v)
            w = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if w > max_w:
                max_w = w
        ws.column_dimensions[col_letter].width = min(max_w + 2, 36)

    print(f'保存 {xlsx}')
    wb.save(xlsx)
    print(f'  sheets: {load_workbook(xlsx, read_only=True).sheetnames}')
    print('完成')


if __name__ == '__main__':
    main()
