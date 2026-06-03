"""按业务合并各店财年报表（全 sheet 合并 + is_test 列 + 储值支付覆盖收款方式）

输入：snowmeet_ai_doc 下 4 业务各店财年报表
  - 租赁 3 店：wanlong / chongli / nanshan
  - 零售 5 店：wanlong / wanlong_service / chongli / nanshan / headquarters
  - 雪票 2 店：chongli / nanshan
  - 养护 3 店：wanlong_service / chongli / nanshan

输出（snowmeet_ai_doc/ 下 4 份）：
  merged_{biz}_orders_fy_2025-05-01_2026-04-30.xlsx

规则：
  - 同 sheet 名纵向拼接；表头取各店联集（按首次出现顺序），缺的列留空
  - 主 sheet「年度{业务}」额外两动作：
    1. 末尾追加 is_test 列（DB [order].is_test 0/1）
    2. 覆盖「收款方式」列：若该订单有任意成功储值支付 → 改写"储值支付"

用法：
  python merge_fy_orders.py            # 默认全跑 4 业务
  python merge_fy_orders.py --biz rent # 只跑某业务
"""
import argparse
import os
import sys
from collections import defaultdict

import pyodbc
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

CONN = ('DRIVER={ODBC Driver 18 for SQL Server};SERVER=tcp:100.28.143.19,1433;'
        'DATABASE=snowmeet_new;UID=claude;PWD=abcd123!@#;'
        'Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=30;')

DOC = r'D:\snowmeet\snowmeet_ai_doc'

INPUTS = {
    'rent': [
        f'{DOC}\\wanlong_rent_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\chongli_rent_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\nanshan_rent_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\huaibei_rent_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\yuyang_rent_orders_fy_2025-05-01_2026-04-30.xlsx',
    ],
    'retail': [
        f'{DOC}\\wanlong_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\wanlong_service_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\chongli_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\nanshan_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\headquarters_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\huaibei_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\yuyang_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
    ],
    'ski_pass': [
        f'{DOC}\\chongli_ski_pass_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\nanshan_ski_pass_orders_fy_2025-05-01_2026-04-30.xlsx',
    ],
    'care': [
        f'{DOC}\\wanlong_service_care_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\chongli_care_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\nanshan_care_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\huaibei_care_orders_fy_2025-05-01_2026-04-30.xlsx',
        f'{DOC}\\yuyang_care_orders_fy_2025-05-01_2026-04-30.xlsx',
    ],
}

MAIN_SHEET = {
    'rent': '年度租赁',
    'retail': '年度零售',
    'ski_pass': '年度雪票',
    'care': '年度养护',
}

HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')


def fetch_db_attrs(codes):
    """返回 {code: (is_test, has_sv_pay)}"""
    result = {}
    if not codes:
        return result
    cn = pyodbc.connect(CONN)
    cu = cn.cursor()
    codes = list({c for c in codes if c})
    BATCH = 900
    sql_tmpl = (
        "SELECT o.code, o.is_test, "
        "CASE WHEN EXISTS ("
        "    SELECT 1 FROM order_payment op "
        "    WHERE op.order_id = o.id "
        "      AND op.status = N'支付成功' AND op.valid = 1 "
        "      AND op.pay_method = N'储值支付'"
        ") THEN 1 ELSE 0 END AS has_sv "
        "FROM [order] o WHERE o.code IN ({})"
    )
    for i in range(0, len(codes), BATCH):
        batch = codes[i:i + BATCH]
        placeholders = ','.join('?' * len(batch))
        cu.execute(sql_tmpl.format(placeholders), batch)
        for r in cu.fetchall():
            # 注：DB 内同 code 可能多条（订单号竞态），取 OR 聚合
            prev = result.get(r.code, (0, 0))
            is_test_v = max(prev[0], int(r.is_test or 0))
            has_sv_v = max(prev[1], int(r.has_sv or 0))
            result[r.code] = (is_test_v, has_sv_v)
    cn.close()
    return result


def read_sheet(path, sheet_name):
    """返回 (headers, rows)；不存在返回 (None, None)"""
    wb = load_workbook(path, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None, None
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        wb.close()
        return [], []
    data = []
    for r in rows_iter:
        row = list(r)
        # 跳过全空行
        if all(v is None or (isinstance(v, str) and v.strip() == '') for v in row):
            continue
        # 对齐到 headers 长度
        if len(row) < len(headers):
            row = row + [None] * (len(headers) - len(row))
        elif len(row) > len(headers):
            row = row[:len(headers)]
        data.append(row)
    wb.close()
    return headers, data


def union_headers(list_of_headers):
    """各表头联集，按首次出现顺序"""
    seen = []
    seen_set = set()
    for hs in list_of_headers:
        for h in hs:
            if h is not None and h not in seen_set:
                seen.append(h)
                seen_set.add(h)
    return seen


def remap_rows(rows, src_headers, dst_headers):
    """按列名重映射；dst 中 src 没有的列填 None"""
    src_idx = {h: i for i, h in enumerate(src_headers) if h is not None}
    out = []
    for row in rows:
        new_row = []
        for h in dst_headers:
            if h in src_idx:
                new_row.append(row[src_idx[h]])
            else:
                new_row.append(None)
        out.append(new_row)
    return out


def visual_len(s):
    """字符视觉宽度：汉字按 2、半角按 1（估算）"""
    if s is None:
        return 0
    return sum(2 if ord(c) > 127 else 1 for c in str(s))


def write_sheet(ws, headers, rows):
    """写表头 + 数据 + 样式 + freeze + 列宽自适应"""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    for r_idx, row in enumerate(rows, 2):
        for c_idx, v in enumerate(row, 1):
            if v is not None:
                ws.cell(r_idx, c_idx, v)
    # 列宽自适应：取前 300 行的最大视觉宽度
    sample_n = min(300, len(rows))
    for c_idx in range(1, len(headers) + 1):
        widths = [visual_len(headers[c_idx - 1])]
        for row in rows[:sample_n]:
            if c_idx - 1 < len(row):
                widths.append(visual_len(row[c_idx - 1]))
        max_w = max(widths) if widths else 8
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_w + 2, 8), 36)
    ws.freeze_panes = 'A2'


def process_biz(biz):
    inputs = INPUTS[biz]
    main_sheet = MAIN_SHEET[biz]
    out_path = f'{DOC}\\merged_{biz}_orders_fy_2025-05-01_2026-04-30.xlsx'

    # 1) 加载所有店所有 sheet
    print(f'\n== 处理业务 {biz} ==')
    sheet_data = defaultdict(list)  # sheet_name → [(path_basename, headers, rows), ...]
    for path in inputs:
        if not os.path.exists(path):
            print(f'  [SKIP] 不存在: {path}')
            continue
        base = os.path.basename(path)
        wb = load_workbook(path, read_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
        for sn in sheets:
            headers, rows = read_sheet(path, sn)
            if headers is not None:
                sheet_data[sn].append((base, headers, rows))
                print(f'  读 {base} [{sn}] = {len(rows)} 行 × {len(headers)} 列')

    # 2) 提取主 sheet 订单号 → DB 一次查
    main_codes = []
    if main_sheet in sheet_data:
        for base, hs, rs in sheet_data[main_sheet]:
            if '订单号' in hs:
                idx = hs.index('订单号')
                for row in rs:
                    code = row[idx]
                    if code is not None and str(code).strip():
                        main_codes.append(str(code).strip())
    db_map = fetch_db_attrs(main_codes)
    print(f'  DB query: {len(set(main_codes))} 个唯一订单号 → {len(db_map)} 命中')

    # 3) 写 xlsx
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    # 排序：主 sheet 第一，其他按出现顺序
    ordered_sheets = list(sheet_data.keys())
    if main_sheet in ordered_sheets:
        ordered_sheets.remove(main_sheet)
        ordered_sheets = [main_sheet] + ordered_sheets

    summary = []
    for sn in ordered_sheets:
        ws = out_wb.create_sheet(sn)
        per_store = sheet_data[sn]
        all_headers = union_headers([hs for _, hs, _ in per_store])

        # 主 sheet 加 is_test 列
        if sn == main_sheet and 'is_test' not in all_headers:
            all_headers = all_headers + ['is_test']

        merged_rows = []
        sv_overwrite_cnt = 0
        is_test_1_cnt = 0
        for base, hs, rs in per_store:
            remapped = remap_rows(rs, hs, all_headers)
            if sn == main_sheet:
                code_idx = all_headers.index('订单号') if '订单号' in all_headers else None
                is_test_idx = all_headers.index('is_test')
                pay_idx = all_headers.index('收款方式') if '收款方式' in all_headers else None
                for row in remapped:
                    code = row[code_idx] if code_idx is not None else None
                    code_str = str(code).strip() if code is not None else ''
                    attr = db_map.get(code_str, (0, 0))
                    row[is_test_idx] = attr[0]
                    if attr[0] == 1:
                        is_test_1_cnt += 1
                    if attr[1] == 1 and pay_idx is not None:
                        row[pay_idx] = '储值支付'
                        sv_overwrite_cnt += 1
            merged_rows.extend(remapped)

        write_sheet(ws, all_headers, merged_rows)
        if sn == main_sheet:
            print(f'  写 sheet「{sn}」: {len(merged_rows)} 行 × {len(all_headers)} 列；'
                  f'is_test=1: {is_test_1_cnt}；收款方式→储值支付: {sv_overwrite_cnt}')
        else:
            print(f'  写 sheet「{sn}」: {len(merged_rows)} 行 × {len(all_headers)} 列')
        summary.append((sn, len(per_store), len(merged_rows), len(all_headers)))

    out_wb.save(out_path)
    print(f'  ✓ 写入 {out_path}')
    return summary


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--biz', choices=list(INPUTS.keys()) + ['all'], default='all')
    args = ap.parse_args()
    targets = list(INPUTS.keys()) if args.biz == 'all' else [args.biz]
    for biz in targets:
        process_biz(biz)


if __name__ == '__main__':
    main()
