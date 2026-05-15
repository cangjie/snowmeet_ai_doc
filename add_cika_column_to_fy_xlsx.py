"""一次性补列：给 wanlong_rent_orders_fy_2025-05-01_2026-04-30.xlsx 加「次卡」列。

规则：
- 该订单存在 rental.valid=1 AND use_card=1 → "是"
- 否则，order_payment.status='支付成功' 笔数 >= 1 → "否"
- 否则（支付成功笔数 = 0） → "-"

用法：python3 add_cika_column_to_fy_xlsx.py
"""
import os
import sys

os.environ.setdefault("ODBCSYSINI", "/opt/homebrew/etc")

import pyodbc  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "wanlong_rent_orders_fy_2025-05-01_2026-04-30.xlsx")
SHEET = "年度租赁"
KEY_HEADER = "订单号"
NEW_HEADER = "次卡"

CONN_STR = (
    "DRIVER={ODBC Driver 18 for SQL Server};"
    "Server=100.28.143.19,1433;Database=snowmeet_new;UID=claude;PWD=abcd123!@#;"
    "Encrypt=Yes;TrustServerCertificate=Yes;"
)

SHOP = "万龙体验中心"
FY_START = "2025-05-01"
FY_END = "2026-05-01"  # exclusive upper bound


def visual_len(s):
    return sum(2 if ord(ch) > 127 else 1 for ch in str(s))


def fetch_dicts():
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()

    cika_codes = set()
    cur.execute(
        """
        SELECT DISTINCT o.code
        FROM [order] o
        JOIN rental r ON r.order_id = o.id
        WHERE r.use_card = 1 AND r.valid = 1
          AND o.shop = ?
          AND o.biz_date >= ? AND o.biz_date < ?
          AND o.code IS NOT NULL AND LTRIM(RTRIM(o.code)) <> ''
        """,
        SHOP, FY_START, FY_END,
    )
    for (code,) in cur.fetchall():
        cika_codes.add(code)

    pay_count = {}
    cur.execute(
        """
        SELECT o.code, COUNT(op.id) AS pay_cnt
        FROM [order] o
        LEFT JOIN order_payment op
               ON op.order_id = o.id AND op.status = N'支付成功'
        WHERE o.shop = ?
          AND o.biz_date >= ? AND o.biz_date < ?
          AND o.code IS NOT NULL AND LTRIM(RTRIM(o.code)) <> ''
        GROUP BY o.code
        """,
        SHOP, FY_START, FY_END,
    )
    for code, cnt in cur.fetchall():
        pay_count[code] = int(cnt or 0)

    conn.close()
    return cika_codes, pay_count


def classify(code, cika_codes, pay_count):
    if code in cika_codes:
        return "是"
    cnt = pay_count.get(code, 0)
    if cnt >= 1:
        return "否"
    return "-"


def main():
    cika_codes, pay_count = fetch_dicts()
    print(f"DB use_card 命中 (订单数): {len(cika_codes)}")
    print(f"DB 订单总数 (有支付聚合): {len(pay_count)}")

    wb = openpyxl.load_workbook(XLSX)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"sheet '{SHEET}' 不存在，实际: {wb.sheetnames}")
    ws = wb[SHEET]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if KEY_HEADER not in headers:
        raise SystemExit(f"表头缺「{KEY_HEADER}」: {headers[:10]}...")
    key_col = headers.index(KEY_HEADER) + 1

    if NEW_HEADER in headers:
        new_col = headers.index(NEW_HEADER) + 1
        print(f"已存在「{NEW_HEADER}」列（第 {new_col} 列），覆盖写入")
    else:
        new_col = ws.max_column + 1
        print(f"新建「{NEW_HEADER}」列（第 {new_col} 列）")

    header_cell = ws.cell(row=1, column=new_col, value=NEW_HEADER)
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.fill = PatternFill("solid", fgColor="1F4E78")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    last_row = ws.max_row
    counts = {"是": 0, "否": 0, "-": 0}
    missing_codes = []
    for r in range(2, last_row + 1):
        code = ws.cell(row=r, column=key_col).value
        if code is None:
            continue
        v = classify(code, cika_codes, pay_count)
        ws.cell(row=r, column=new_col, value=v).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        counts[v] += 1
        if v != "是" and code not in pay_count:
            missing_codes.append((r, code))

    max_len = visual_len(NEW_HEADER)
    for v in counts:
        if visual_len(v) > max_len:
            max_len = visual_len(v)
    ws.column_dimensions[get_column_letter(new_col)].width = min(max_len + 4, 36)

    wb.save(XLSX)

    size_kb = os.path.getsize(XLSX) / 1024
    print(f"\nxlsx 行数: {last_row - 1}")
    print(f"  是: {counts['是']}")
    print(f"  否: {counts['否']}")
    print(f"  -:  {counts['-']}")
    print(f"  合计: {sum(counts.values())}")
    if missing_codes:
        print(f"  ⚠ 在 DB 聚合中找不到的订单号: {len(missing_codes)} 条")
        for r, c in missing_codes[:5]:
            print(f"    行 {r} 订单号 {c}")
    print(f"\n写入 {XLSX}\n文件大小: {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
