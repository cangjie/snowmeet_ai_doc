"""把 6-1 git 版本的「七色米订单号」列回填到新生成的 fy xlsx。

5 个 retail xlsx 各取 git 14f32e0 commit 的「年度零售」sheet，
按 订单号 → 七色米订单号 建 dict，覆盖到新文件末尾追加列。
"""
import os
import subprocess
import sys
import tempfile

sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

DOC = 'D:/snowmeet/snowmeet_ai_doc'
COMMIT = '14f32e0'
SHEET_NAME = '年度零售'
MI7_COL = '七色米订单号'

FILES = [
    'nanshan_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
    'chongli_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
    'wanlong_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
    'wanlong_service_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
    'headquarters_retail_orders_fy_2025-05-01_2026-04-30.xlsx',
]


def extract_mi7_map(fname):
    """从 git COMMIT 版的 xlsx 提取 code → mi7 dict"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tf:
        tmp = tf.name
    try:
        r = subprocess.run(['git', '-C', DOC, 'show', f'{COMMIT}:{fname}'],
                           capture_output=True)
        if r.returncode != 0:
            print(f'  git show 失败: {r.stderr.decode("utf-8", "replace")}')
            return {}
        with open(tmp, 'wb') as f:
            f.write(r.stdout)
        wb = load_workbook(tmp, read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            print(f'  老版 sheet「{SHEET_NAME}」不存在')
            return {}
        ws = wb[SHEET_NAME]
        hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if MI7_COL not in hdrs or '订单号' not in hdrs:
            print(f'  老版缺列 mi7 in={MI7_COL in hdrs} 订单号 in={"订单号" in hdrs}')
            return {}
        code_i = hdrs.index('订单号')
        mi7_i = hdrs.index(MI7_COL)
        mp = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[code_i]
            mi7 = row[mi7_i]
            if code:
                key = str(code).strip()
                # 同 code 可能多行（去重前），保留有 mi7 的
                if key not in mp or (mi7 and not mp[key]):
                    mp[key] = mi7
        wb.close()
        return mp
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass


def add_mi7_col(fname, mp):
    """在新 xlsx 的 年度零售 sheet 末尾追加 七色米订单号 列"""
    path = f'{DOC}/{fname}'
    wb = load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        print(f'  新版 sheet「{SHEET_NAME}」不存在')
        wb.close()
        return 0
    ws = wb[SHEET_NAME]
    n_cols = ws.max_column
    n_rows = ws.max_row
    hdrs = [ws.cell(1, c).value for c in range(1, n_cols + 1)]
    if MI7_COL in hdrs:
        # 已有此列，更新值即可
        col_idx = hdrs.index(MI7_COL) + 1
    else:
        col_idx = n_cols + 1
        # 写表头（拷贝相邻列样式）
        ref_cell = ws.cell(1, n_cols)
        new_cell = ws.cell(1, col_idx, MI7_COL)
        # 蓝底白字粗体
        new_cell.fill = PatternFill('solid', fgColor='1F4E78')
        new_cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        new_cell.alignment = Alignment(horizontal='center', vertical='center')

    code_idx = hdrs.index('订单号') + 1 if '订单号' in hdrs else None
    if code_idx is None:
        print('  新版主 sheet 缺「订单号」列')
        wb.close()
        return 0
    filled = 0
    for r in range(2, n_rows + 1):
        code = ws.cell(r, code_idx).value
        if code is None:
            continue
        key = str(code).strip()
        if key in mp and mp[key]:
            ws.cell(r, col_idx, mp[key])
            filled += 1
    wb.save(path)
    wb.close()
    return filled


for f in FILES:
    print(f'\n[{f}]')
    mp = extract_mi7_map(f)
    print(f'  老版 code→mi7 dict: {len(mp)} 项（{sum(1 for v in mp.values() if v)} 项有 mi7）')
    if not mp:
        continue
    n = add_mi7_col(f, mp)
    print(f'  回填到新版: {n} 行有 mi7')
